import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import Tk, messagebox
import getpass
 
 
# ---------------- CONFIG ----------------
INPUT_FILE = r"C:\newproj\OPE_Samples.xlsx"          # uploaded Excel
DATABASE_FILE = r"C:\newproj\Book1.xlsx"      # main database Excel
DAILY_LIMIT = 200
 
# Dynamically capture uploader (Windows login)
UPLOADED_BY = getpass.getuser()
 
print(f"👤 Uploaded By (system user): {UPLOADED_BY}")
 
# ---------------- READ EXCEL ----------------
df = pd.read_excel(INPUT_FILE)
df.columns = df.columns.str.strip()
 
print("\n📌 Columns found in Excel:")
print(list(df.columns))
 
# ---------------- AUTO-DETECT COLUMNS ----------------
amount_col = next((c for c in df.columns if "amount" in c.lower()), None)
date_col = next((c for c in df.columns if "date" in c.lower()), None)
employee_col = next(
    (c for c in df.columns if "employee" in c.lower() and "code" in c.lower()),
    None
)
 
print("\n📌 Auto-detected columns:")
print(f"   Employee Column : {employee_col}")
print(f"   Date Column     : {date_col}")
print(f"   Amount Column   : {amount_col}")
 
if not all([amount_col, date_col, employee_col]):
    raise Exception("❌ Required columns not found in Excel")
 
# ---------------- DATA CLEANING ----------------
df[amount_col] = pd.to_numeric(df[amount_col], errors="coerce").fillna(0)
df[date_col] = pd.to_datetime(df[date_col], errors="coerce", dayfirst=True)
 
print("\n🧹 Data cleaned successfully")
 
# ---------------- DAILY TOTAL & VALIDATION ----------------
daily_total = df.groupby([employee_col, date_col])[amount_col].transform("sum")
 
df["Validation"] = np.where(
    (daily_total > DAILY_LIMIT) | (df[amount_col] > DAILY_LIMIT),
    "Exceeded daily limit",
    "Within limit"
)
 
df["UploadedBy"] = UPLOADED_BY
 
print("\n✅ Validation completed")
print(df[[employee_col, date_col, amount_col, "Validation"]].head())
 
# ---------------- MERGE INTO DATABASE ----------------
if os.path.exists(DATABASE_FILE):
    old_df = pd.read_excel(DATABASE_FILE)
    final_df = pd.concat([old_df, df], ignore_index=True)
else:
    final_df = df
 
final_df.to_excel(DATABASE_FILE, index=False)
 
print(f"\n💾 Data saved to: {DATABASE_FILE}")
 
# ---------------- HIGHLIGHT EXCEEDED ----------------
wb = load_workbook(DATABASE_FILE)
ws = wb.active
 
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
 
validation_col = next(
    (i for i, c in enumerate(ws[1], start=1) if c.value == "Validation"),
    None
)
 
for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=validation_col).value == "Exceeded daily limit":
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = red_fill
 
wb.save(DATABASE_FILE)
 
print("\n🎨 Highlighted exceeded rows in Excel")
 
# ---------------- POPUP ----------------
root = Tk()
root.withdraw()
 
if (df["Validation"] == "Exceeded daily limit").any():
    messagebox.showwarning(
        "Daily Limit Exceeded",
        f"One or more employees exceeded ₹{DAILY_LIMIT}"
    )